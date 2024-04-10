from openpyxl import load_workbook
import pandas as pd

my_path = 'C:\\Users\\l.haynes\\Desktop\\py_edit\\files\\'

wb = load_workbook(f"{my_path}test.xlsx")
ws = wb['Sheet1']


df = pd.DataFrame(ws.values)

print(df)

num1 = ws['B2'].value
num2 = ws['B3'].value

ws['B2'].value = num1 * 2
ws['B3'].value = num2 * 2

wb.save(f"{my_path}test.xlsx")

# print(ws)
# print(num2)
print('done')




